from flask import Flask, render_template, Response, jsonify
import cv2
import torch
import easyocr
import numpy as np
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage

app = Flask(__name__)

class Parameters:
    def __init__(self):
        self.rect_size = 0.5
        self.region_threshold = 0.5

params = Parameters()


text_reader = easyocr.Reader(['en'])

camera = cv2.VideoCapture(0)

def load_yolov5_model(model_path, device):
    model = torch.hub.load('ultralytics/yolov5', 'custom', path=model_path, device=device)
    return model

import torch

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

model_anpr = load_yolov5_model(r'model\best.pt', device)
model_face = load_yolov5_model(r'yolov5m.pt', device)

def detection(frame, model):
    results = model(frame)
    detected_objects = []
    for result in results.xyxy[0]:
        confidence = result[4].item()
        class_id = int(result[5].item())
        if confidence > 0.5 and class_id == 0:
            x1, y1, x2, y2 = result[:4]
            detected_objects.append((x1, y1, x2-x1, y2-y1))
    return detected_objects

def save_cropped_image(image, box, directory, prefix="crop"):
    if not os.path.exists(directory):
        os.makedirs(directory)
    x, y, w, h = box
    crop_img = image[int(y):int(y+h), int(x):int(x+w)]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S%f")
    file_path = os.path.join(directory, f"{prefix}_{timestamp}.jpg")
    cv2.imwrite(file_path, crop_img)
    return file_path

def insert_images_to_excel(excel_path, number_plate_paths, face_paths, headers, ocr_texts=[]):
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
    else:
        wb = load_workbook(excel_path)
        ws = wb.active
    
    for i, (number_plate_path, face_path) in enumerate(zip(number_plate_paths, face_paths)):
        row = [datetime.now()]
        number_plate_img = PILImage.open(number_plate_path)
        number_plate_img.thumbnail((100, 100), PILImage.LANCZOS)
        number_plate_thumb_path = number_plate_path.replace('.jpg', '_thumb.jpg')
        number_plate_img.save(number_plate_thumb_path)
        number_plate_excel_img = ExcelImage(number_plate_thumb_path)
        
        face_img = PILImage.open(face_path)
        face_img.thumbnail((100, 100), PILImage.LANCZOS)
        face_thumb_path = face_path.replace('.jpg', '_thumb.jpg')
        face_img.save(face_thumb_path)
        face_excel_img = ExcelImage(face_thumb_path)
        
        ws.append(row)
        row_index = ws.max_row
        ws.add_image(number_plate_excel_img, f'B{row_index}')
        ws.add_image(face_excel_img, f'C{row_index}')
        if ocr_texts:
            ws[f'D{row_index}'] = ocr_texts[i] if i < len(ocr_texts) else ""
    
    wb.save(excel_path)

def filter_text(rect_size, ocr_results, region_threshold):
    text = []
    for result in ocr_results:
        if result[2] > region_threshold:
            text.append(result[1])
    return text

def gen_frames():
    while True:
        success, frame = camera.read()
        if not success:
            break
        else:
            detected_anpr = detection(frame, model_anpr)
            detected_face = detection(frame, model_face)

            detected = frame.copy()
            for x, y, w, h in detected_anpr + detected_face:
                cv2.rectangle(detected, (int(x), int(y)), (int(x+w), int(y+h)), (0, 255, 0), 2)

            number_plate_paths = [save_cropped_image(frame, box, "Cropped_Number_Plates", "number_plate") for box in detected_anpr]
            face_paths = [save_cropped_image(frame, box, "Cropped_Faces", "face") for box in detected_face]

            if number_plate_paths and face_paths:
                insert_images_to_excel("Detection_Images/detections.xlsx", number_plate_paths, face_paths, ["Timestamp", "Number Plate", "Face", "OCR Text"])

            ret, buffer = cv2.imencode('.jpg', detected)
            frame = buffer.tobytes()

            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

@app.route('/video_feed')
def video_feed():
    return Response(gen_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/ocr_results')
def ocr_results():
    success, frame = camera.read()
    if not success:
        return jsonify([])

    detected_anpr = detection(frame, model_anpr)
    text_results = []

    for box in detected_anpr:
        x, y, w, h = box
        crop_img = frame[int(y):int(y+h), int(x):int(x+w)]
        result_easyocr = text_reader.readtext(crop_img)
        text = filter_text(params.rect_size, result_easyocr, params.region_threshold)
        if text:
            text_results.append(text[-1])

    number_plate_paths = [save_cropped_image(frame, box, "Cropped_Number_Plates", "number_plate") for box in detected_anpr]
    face_paths = [save_cropped_image(frame, box, "Cropped_Faces", "face") for box in detection(frame, model_face)]

    if number_plate_paths and face_paths:
        insert_images_to_excel("Detection_Images/detections.xlsx", number_plate_paths, face_paths, ["Timestamp", "Number Plate", "Face", "OCR Text"], text_results)

    return jsonify(text_results)

@app.route('/')
def index():
    return render_template('index3.html')

if __name__ == "__main__":
    app.run(debug=True)
