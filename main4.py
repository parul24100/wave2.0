from flask import Flask, render_template, Response
import cv2
import torch
import easyocr
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage

app = Flask(__name__)

class Parameters:
    def __init__(self):
        self.rect_size = 0.5
        self.region_threshold = 0.5

def load_yolov5_model(model_path, device):
    model = torch.hub.load('ultralytics/yolov5', 'custom', path=model_path, device=device)
    return model

def detection(frame, model):
    results = model(frame)
    detected_objects = []
    for result in results.xyxy[0]:
        confidence = result[4].item()
        class_id = int(result[5].item())
        if confidence > 0.5 and class_id == 0:
            x1, y1, x2, y2 = result[:4]
            detected_objects.append((x1, y1, x2-x1, y2-y1))
    return detected_objects

def filter_text(rect_size, ocr_results, region_threshold):
    text = []
    for result in ocr_results:
        if result[2] > region_threshold:
            text.append(result[1])
    return text

def save_results(text, filename, directory):
    if not os.path.exists(directory):
        os.makedirs(directory)
    with open(os.path.join(directory, filename), 'a') as f:
        f.write(f"{datetime.now()},{text}\n")

def save_cropped_image(image, box, directory, prefix="crop"):
    if not os.path.exists(directory):
        os.makedirs(directory)
    x, y, w, h = box
    crop_img = image[int(y):int(y+h), int(x):int(x+w)]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S%f")
    file_path = os.path.join(directory, f"{prefix}_{timestamp}.jpg")
    cv2.imwrite(file_path, crop_img)
    return file_path

def insert_images_to_excel(excel_path, number_plate_paths, face_paths, headers):
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
    else:
        wb = load_workbook(excel_path)
        ws = wb.active
    
    for number_plate_path, face_path in zip(number_plate_paths, face_paths):
        row = [datetime.now()]
        number_plate_img = PILImage.open(number_plate_path)
        number_plate_img.thumbnail((100, 100), PILImage.LANCZOS)
        number_plate_thumb_path = number_plate_path.replace('.jpg', '_thumb.jpg')
        number_plate_img.save(number_plate_thumb_path)
        number_plate_excel_img = ExcelImage(number_plate_thumb_path)
        
        face_img = PILImage.open(face_path)
        face_img.thumbnail((100, 100), PILImage.LANCZOS)
        face_thumb_path = face_path.replace('.jpg', '_thumb.jpg')
        face_img.save(face_thumb_path)
        face_excel_img = ExcelImage(face_thumb_path)
        
        ws.append(row)
        row_index = ws.max_row
        ws.add_image(number_plate_excel_img, f'B{row_index}')
        ws.add_image(face_excel_img, f'C{row_index}')
    
    wb.save(excel_path)

params = Parameters()

import torch

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

model_anpr = load_yolov5_model(r'C:\Users\prajw\Downloads\Bagalkot Hackathon\New folder (3)\Automatic_Number_Plate_Recognition_YOLO_OCR\model\best.pt', device)
model_face = load_yolov5_model(r'C:\Users\prajw\Downloads\Bagalkot Hackathon\New folder (3)\Automatic_Number_Plate_Recognition_YOLO_OCR\yolov5m.pt', device)
text_reader = easyocr.Reader(['en'])

def generate_frames():
    camera = cv2.VideoCapture(0)
    if not camera.isOpened():
        print("Error: Could not open video capture device.")
        return

    excel_path = "Detection_Images/detections.xlsx"
    headers = ["Timestamp", "Number Plate", "Face"]

    while True:
        success, frame = camera.read()
        if not success:
            break
        else:
            try:
                detected_anpr = detection(frame, model_anpr)
                detected_face = detection(frame, model_face)
                detected = frame.copy()
                for x, y, w, h in detected_anpr + detected_face:
                    cv2.rectangle(detected, (int(x), int(y)), (int(x+w), int(y+h)), (0, 255, 0), 2)

                number_plate_paths = [save_cropped_image(frame, box, "Cropped_Number_Plates", "number_plate") for box in detected_anpr]
                face_paths = [save_cropped_image(frame, box, "Cropped_Faces", "face") for box in detected_face]

                if number_plate_paths and face_paths:
                    insert_images_to_excel(excel_path, number_plate_paths, face_paths, headers)

                resulteasyocr = text_reader.readtext(detected)
                text = filter_text(params.rect_size, resulteasyocr, params.region_threshold)
                if text:
                    save_results(text[-1], "ocr_results.csv", "Detection_Images")

                ret, buffer = cv2.imencode('.jpg', detected)
                frame = buffer.tobytes()
                yield (b'--frame\r\n'
                       b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

            except Exception as e:
                print(f"Error processing frame: {e}")
                continue

    camera.release()

@app.route('/')
def index():
    return render_template('index2.html')

@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

if __name__ == "__main__":
    app.run(debug=True)
